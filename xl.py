import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def actualizarxl(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    # cell=sheet["a1"] o cell=sheet.cell(1,1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        presio_correcto = cell.value * 2
        presio_correcto_celda = sheet.cell(row, 4)
        presio_correcto_celda.value = presio_correcto

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e1")

    wb.save(filename)
